#!/usr/bin/env python3
"""
Convert a Hoang Thanh daily-sales-journal sheet (NHAT KY BAN HANG HANG NGAY)
into one flat CSV per day, ready for Demo_CSV's generic column mapper.

This runs OUTSIDE the app. Demo_CSV's uploader reads a flat table with one
header row and writes every row of a file against a single date, so a whole
month cannot go in as one file -- it would silently collapse into one day
holding the month's sum. Splitting here keeps the app's upload pipeline
untouched.

Usage:
    python tools/convert_journal.py Book1.xlsx -o out/
    python tools/convert_journal.py Book1.xlsx -o out/ --sheet "T02 (2)"

Source layout this understands (only the six focus columns are read):

    A   CA           shift label
    B   NGAY THANG   date -- rows whose value is "TONG" are subtotals, dropped
    O   TEN HANG     product, e.g. "A95 (TRU 2)" or "N50"
    R   SB           volume sold
    S   DG           unit price
    T   TT           amount

Everything else in the sheet is deliberately ignored.
"""
import argparse
import csv
import datetime
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# Sheet is 1-indexed; these are the only columns read.
COL_CA, COL_DATE, COL_NAME, COL_SB, COL_DG, COL_TT = 1, 2, 15, 18, 19, 20
FIRST_DATA_ROW = 7

PUMP_RE = re.compile(r"^(.*?)\s*\(\s*TR[UỤ]\s*(\d+)\s*\)$", re.IGNORECASE)

# Shift labels differing only by vowel shape (e.g. a name spelled with "ă" vs
# one spelled with "â") are DIFFERENT shifts.
# Naive ASCII folding turns both into "van" and silently merges them, so the
# vowel shape has to survive into the slug. Telex spelling does that and is
# recognisable to a Vietnamese reader. Tone marks carry no distinction we need
# here and are dropped.
#
#   ă -> aw   â -> aa   ê -> ee   ô -> oo   ơ -> ow   ư -> uw   đ -> dd
SHAPE = {
    ("a", "̆"): "aw",   # ă  breve
    ("a", "̂"): "aa",   # â  circumflex
    ("e", "̂"): "ee",   # ê
    ("o", "̂"): "oo",   # ô
    ("o", "̛"): "ow",   # ơ  horn
    ("u", "̛"): "uw",   # ư
}
TONE_MARKS = {"̀", "́", "̃", "̉", "̣"}


def slug(text):
    """Lowercase ASCII slug that preserves Vietnamese vowel-shape distinctions.

    Must not fold ă/â (or ơ/ư, ê/e, ô/o) together: two shifts or products whose
    names differ only by vowel shape would merge into one metric and their data
    would be summed as if they were the same thing.
    """
    s = unicodedata.normalize("NFD", str(text).lower().replace("đ", "dd"))
    out, i = [], 0
    while i < len(s):
        ch = s[i]
        if unicodedata.category(ch) == "Mn":   # stray combining mark
            i += 1
            continue
        marks, j = [], i + 1
        while j < len(s) and unicodedata.category(s[j]) == "Mn":
            marks.append(s[j])
            j += 1
        shape = next((m for m in marks if m not in TONE_MARKS), None)
        out.append(SHAPE.get((ch, shape), ch) if shape else ch)
        i = j
    return re.sub(r"[^a-z0-9]+", "_", "".join(out)).strip("_")


def num(v):
    return v if isinstance(v, (int, float)) else 0.0


def read_rows(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows, dropped = [], 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        date_cell = ws.cell(r, COL_DATE).value
        # Subtotal rows carry "TONG" in the date column -- drop the whole row.
        if isinstance(date_cell, str) and "TỔNG" in date_cell.upper():
            dropped += 1
            continue
        name = ws.cell(r, COL_NAME).value
        if not name or not isinstance(date_cell, datetime.datetime):
            continue
        raw = str(name).strip()
        m = PUMP_RE.match(raw)
        rows.append({
            "ca": str(ws.cell(r, COL_CA).value).strip() if ws.cell(r, COL_CA).value else "",
            "date": date_cell.date(),
            "fuel": m.group(1).strip().upper() if m else None,
            "pump": int(m.group(2)) if m else None,
            "item": None if m else raw,
            "sb": num(ws.cell(r, COL_SB).value),
            "dg": num(ws.cell(r, COL_DG).value),
            "tt": num(ws.cell(r, COL_TT).value),
        })
    return rows, dropped, (sheet_name or wb.sheetnames[0])


def build(rows):
    """Collapse to one record per day. A day may hold two price blocks (the
    weekly Thursday adjustment); those are summed, and the day's unit price
    becomes the volume-weighted average of both blocks."""
    pumps = sorted({(r["fuel"], r["pump"]) for r in rows if r["pump"]}, key=lambda p: (p[0], p[1]))
    items = sorted({r["item"] for r in rows if r["item"]})
    shifts = sorted({r["ca"] for r in rows if r["ca"]})
    fuels = sorted({f for f, _ in pumps})

    days = defaultdict(lambda: {
        "pump_sb": defaultdict(float), "pump_tt": defaultdict(float),
        "item_sb": defaultdict(float), "item_tt": defaultdict(float),
        "fuel_sb": defaultdict(float), "fuel_tt": defaultdict(float),
        "shifts": set(),
    })
    for r in rows:
        d = days[r["date"]]
        if r["ca"]:
            d["shifts"].add(r["ca"])
        if r["pump"]:
            key = (r["fuel"], r["pump"])
            d["pump_sb"][key] += r["sb"]
            d["pump_tt"][key] += r["tt"]
            d["fuel_sb"][r["fuel"]] += r["sb"]
            d["fuel_tt"][r["fuel"]] += r["tt"]
        else:
            d["item_sb"][r["item"]] += r["sb"]
            d["item_tt"][r["item"]] += r["tt"]

    header = ["date"]
    header += [f"ca_{slug(s)}" for s in shifts]
    for f, p in pumps:
        header += [f"sl_{slug(f)}_tru{p}", f"tt_{slug(f)}_tru{p}"]
    header += [f"gia_{slug(f)}" for f in fuels]
    for it in items:
        header += [f"sl_{slug(it)}", f"tt_{slug(it)}"]

    # Two different source names collapsing to one column would silently sum
    # unrelated data. Fail loudly instead -- this is the failure mode that is
    # invisible once it reaches the dashboard.
    dupes = {h for h in header if header.count(h) > 1}
    if dupes:
        raise SystemExit(
            "Column name collision: " + ", ".join(sorted(dupes)) +
            "\nTwo distinct source names produced the same slug. Fix slug() "
            "before continuing -- their data would otherwise be merged."
        )

    records = []
    for date in sorted(days):
        d = days[date]
        row = {"date": date.strftime("%d/%m/%Y")}
        for s in shifts:
            row[f"ca_{slug(s)}"] = 1 if s in d["shifts"] else 0
        for f, p in pumps:
            row[f"sl_{slug(f)}_tru{p}"] = round(d["pump_sb"][(f, p)], 3)
            row[f"tt_{slug(f)}_tru{p}"] = round(d["pump_tt"][(f, p)], 2)
        for f in fuels:
            vol = d["fuel_sb"][f]
            # Blank, not zero, when nothing sold: an absent price is not a
            # price of 0, and a 0 would drag the price chart to the floor.
            row[f"gia_{slug(f)}"] = round(d["fuel_tt"][f] / vol, 2) if vol else ""
        for it in items:
            row[f"sl_{slug(it)}"] = round(d["item_sb"][it], 3)
            row[f"tt_{slug(it)}"] = round(d["item_tt"][it], 2)
        records.append((date, row))
    return header, records, {"pumps": pumps, "items": items, "shifts": shifts}


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("-o", "--outdir", default="out")
    ap.add_argument("--sheet", default=None, help="sheet name (default: first sheet)")
    ap.add_argument("--prefix", default="nhat-ky", help="output filename prefix")
    args = ap.parse_args()

    rows, dropped, sheet = read_rows(args.workbook, args.sheet)
    if not rows:
        sys.exit(f"No usable rows found in sheet {sheet!r}.")
    header, records, meta = build(rows)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    for date, row in records:
        # Date in the filename so the app can recover it even if the date
        # column is remapped.
        path = outdir / f"{args.prefix}-{date.isoformat()}.csv"
        with path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=header)
            w.writeheader()
            w.writerow(row)

    total_tt = sum(r["tt"] for r in rows)
    fuel_sb = sum(r["sb"] for r in rows if r["pump"])
    print(f"sheet          : {sheet}")
    print(f"rows read      : {len(rows)}   (dropped {dropped} TONG subtotal rows)")
    print(f"days written   : {len(records)} -> {outdir}/")
    print(f"columns        : {len(header)}")
    print(f"shifts         : {', '.join(f'{s} -> ca_{slug(s)}' for s in meta['shifts'])}")
    print(f"pumps          : {', '.join(f'{f} tru {p}' for f, p in meta['pumps'])}")
    print(f"other products : {', '.join(meta['items']) or '(none)'}")
    print(f"total amount   : {total_tt:,.0f}")
    print(f"fuel volume    : {fuel_sb:,.2f}")


if __name__ == "__main__":
    main()
